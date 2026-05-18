"""Сервис агрегированных отчетов и экспорта.

Агрегаты для дашборда замполита:
  - средние/медианные метрики по отделу/участку/должности;
  - распределение anchor_engine с подсветкой "anchor_pretender";
  - топ-ключевых слов «образа компании» в свободном тексте;
  - тренды по циклам.
"""

from __future__ import annotations

import io
from collections import Counter
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.analysis import AnalysisResult
from app.models.response import TestResponse
from app.models.test import Test
from app.models.user import User

METRIC_KEYS = [
    "stress_index",
    "burnout_risk",
    "work_life_balance",
    "employer_brand",
    "leadership_trust",
    "safety_culture",
    "career_clarity",
    "team_cohesion",
    "loyalty_intent",
    "psychological_safety",
]


def _mean(values: list[float]) -> float:
    return round(sum(values) / len(values), 2) if values else 0.0


def _company_phase(
    company_metrics: dict[str, float],
    anchor_counter: Counter[str],
) -> dict[str, Any]:
    """Сводная «фаза компании»: разбитие / среднее / стагнация.

    Эвристика:
      - engines_share — доля «двигателей» среди классифицированных;
      - pretenders_share — доля «якорей под маской двигателей»;
      - composite — среднее ключевых позитивных метрик.

    Пороги выбраны мягко: при пустой/малой выборке падаем в «среднее».
    """
    total_classified = sum(anchor_counter.values())
    engines = anchor_counter.get("engine", 0)
    anchors = anchor_counter.get("anchor", 0) + anchor_counter.get("anchor_pretender", 0)
    engines_share = engines / total_classified if total_classified else 0.0
    anchors_share = anchors / total_classified if total_classified else 0.0

    positive = (
        company_metrics.get("employer_brand", 0)
        + company_metrics.get("leadership_trust", 0)
        + company_metrics.get("loyalty_intent", 0)
        + company_metrics.get("team_cohesion", 0)
    ) / 4 if company_metrics else 0.0

    if total_classified == 0:
        phase = "среднее"
        reason = "Данных пока недостаточно — собираем."
    elif engines_share >= 0.45 and positive >= 3.8 and anchors_share <= 0.15:
        phase = "разбитие"
        reason = (
            f"Двигателей {round(engines_share * 100)}%, "
            f"якорей {round(anchors_share * 100)}%, "
            f"среднее по позитивным шкалам {positive:.1f}/5."
        )
    elif anchors_share >= 0.35 or positive < 3.0:
        phase = "стагнация"
        reason = (
            f"Якорей {round(anchors_share * 100)}%, "
            f"среднее по позитивным шкалам {positive:.1f}/5. "
            "Команда теряет темп."
        )
    else:
        phase = "среднее"
        reason = (
            f"Двигателей {round(engines_share * 100)}%, "
            f"якорей {round(anchors_share * 100)}%, "
            f"среднее по позитивным шкалам {positive:.1f}/5."
        )

    return {
        "phase": phase,
        "reason": reason,
        "engines_share": round(engines_share, 2),
        "anchors_share": round(anchors_share, 2),
        "composite_positive": round(positive, 2),
    }


async def build_dashboard_payload(
    db: AsyncSession, *, cycle_tag: str | None = None
) -> dict[str, Any]:
    stmt = (
        select(AnalysisResult, TestResponse, User)
        .join(TestResponse, AnalysisResult.response_id == TestResponse.id)
        .join(User, TestResponse.user_id == User.id)
    )
    if cycle_tag:
        stmt = stmt.where(TestResponse.cycle_tag == cycle_tag)

    rows = (await db.execute(stmt)).all()

    # Счётчик прохождений по тестам: используем все ответы (даже без
    # анализа), чтобы было видно реальный охват.
    from sqlalchemy import func

    test_stats_stmt = (
        select(
            Test.id,
            Test.title,
            func.count(TestResponse.id).label("submissions"),
            func.count(func.distinct(TestResponse.respondent_name)).label("unique_names"),
        )
        .outerjoin(TestResponse, TestResponse.test_id == Test.id)
        .group_by(Test.id, Test.title)
        .order_by(Test.title)
    )
    if cycle_tag:
        test_stats_stmt = test_stats_stmt.where(
            (TestResponse.cycle_tag == cycle_tag) | (TestResponse.id.is_(None))
        )
    test_stats_rows = (await db.execute(test_stats_stmt)).all()
    test_stats = [
        {
            "test_id": str(r.id),
            "title": r.title,
            "submissions": int(r.submissions or 0),
            "unique_respondents": int(r.unique_names or 0),
        }
        for r in test_stats_rows
    ]
    total_submissions = sum(s["submissions"] for s in test_stats)
    total_unique = (
        await db.execute(
            select(func.count(func.distinct(TestResponse.respondent_name)))
        )
    ).scalar() or 0

    totals: dict[str, list[float]] = {k: [] for k in METRIC_KEYS}
    by_department: dict[str, dict[str, list[float]]] = {}
    by_site: dict[str, dict[str, list[float]]] = {}
    anchor_counter: Counter[str] = Counter()
    keywords_counter: Counter[str] = Counter()
    image_sentences: list[str] = []
    anchor_pretenders: list[dict[str, Any]] = []
    alerts: list[dict[str, Any]] = []

    for analysis, response, user in rows:
        for k in METRIC_KEYS:
            v = analysis.score_metrics.get(k)
            if v is None:
                continue
            totals[k].append(float(v))
            dept_bucket = by_department.setdefault(
                user.department or "—", {k: [] for k in METRIC_KEYS}
            )
            dept_bucket[k].append(float(v))
            site_bucket = by_site.setdefault(
                user.site or "—", {k: [] for k in METRIC_KEYS}
            )
            site_bucket[k].append(float(v))

        anchor_counter[analysis.anchor_engine or "unknown"] += 1

        ei = analysis.employer_image or {}
        for kw in ei.get("keywords", []) or []:
            keywords_counter[kw.lower().strip()] += 1
        if sentence := ei.get("one_sentence"):
            image_sentences.append(sentence)

        if analysis.anchor_engine == "anchor_pretender":
            anchor_pretenders.append(
                {
                    "user_id": str(user.id),
                    "department": user.department,
                    "site": user.site,
                    "position": user.position,
                    "confidence": analysis.anchor_engine_confidence,
                    "reasoning": analysis.anchor_engine_reasoning,
                    "cycle_tag": response.cycle_tag,
                }
            )

        for flag in analysis.risk_flags or []:
            alerts.append(
                {
                    "code": flag,
                    "department": user.department,
                    "site": user.site,
                    "anchor_engine": analysis.anchor_engine,
                }
            )

    company_metrics = {k: _mean(v) for k, v in totals.items()}
    phase = _company_phase(company_metrics, anchor_counter)

    return {
        "cycle_tag": cycle_tag or "all",
        "respondents": len(rows),
        "company_phase": phase,
        "company_metrics": company_metrics,
        "by_department": {
            d: {k: _mean(v) for k, v in metrics.items()}
            for d, metrics in by_department.items()
        },
        "by_site": {
            s: {k: _mean(v) for k, v in metrics.items()}
            for s, metrics in by_site.items()
        },
        "anchor_engine_distribution": dict(anchor_counter),
        "anchor_pretenders": anchor_pretenders,
        "employer_image": {
            "top_keywords": [
                {"keyword": k, "weight": w}
                for k, w in keywords_counter.most_common(15)
            ],
            "sample_sentences": image_sentences[:8],
        },
        "alerts_summary": Counter(a["code"] for a in alerts).most_common(),
        "test_stats": test_stats,
        "total_submissions": total_submissions,
        "total_unique_respondents": int(total_unique),
    }


def export_excel(dashboard: dict[str, Any]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Company"
    ws.append(["Метрика", "Среднее (0..5)"])
    for k, v in dashboard["company_metrics"].items():
        ws.append([k, v])

    ws_d = wb.create_sheet("By department")
    ws_d.append(["Отдел", *METRIC_KEYS])
    for dept, metrics in dashboard["by_department"].items():
        ws_d.append([dept, *(metrics[k] for k in METRIC_KEYS)])

    ws_a = wb.create_sheet("Anchor pretenders")
    ws_a.append(
        ["user_id", "department", "site", "position", "confidence", "reasoning"]
    )
    for row in dashboard["anchor_pretenders"]:
        ws_a.append(
            [
                row["user_id"],
                row.get("department"),
                row.get("site"),
                row.get("position"),
                row.get("confidence"),
                row.get("reasoning"),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(dashboard: dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 50
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "BuildPulse — отчёт по компании")
    y -= 30
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Цикл: {dashboard['cycle_tag']}    Респондентов: {dashboard['respondents']}")
    y -= 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Средние метрики")
    y -= 18
    c.setFont("Helvetica", 10)
    for k, v in dashboard["company_metrics"].items():
        c.drawString(60, y, f"{k}: {v}")
        y -= 14
        if y < 60:
            c.showPage()
            y = height - 50
    c.showPage()
    c.save()
    return buf.getvalue()
